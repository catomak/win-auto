# from typing import override
from pywinauto.application import Application, AppStartError, ProcessNotFoundError
from pywinauto.findbestmatch import MatchError
from time import sleep
import openpyxl
import re
from service import Helper, config, log, ERRORS


class AppWorker:
    """
    Class for the working with windows applications
    """

    def __init__(self):
        self.program_name = self.__class__.__name__.replace('Worker', '')
        self.app_conf = config[self.program_name]
        self.program_obj = self.launch(self.app_conf['program_path'], self.app_conf['launch_type'])
        self.executed = False

    @staticmethod
    def connect(program: str):
        """Method for connecting to an opened program"""

        try:
            return Application(backend='uia').connect(path=program)
        except FileNotFoundError:
            log.exception(f'Fail connection to  {program}')

    def launch(self, program_path: str, launch_type: str = 'normal'):
        """
        Function for launch a program using one of two method
        :param program_path: the full program path
        :param launch_type: if value == 'normal' an application will be started by calling the executable file
                            if value == 'manual' an application will be started by double-click on program folder
        :return: the instance of Application class type
        """

        launch_types = {
            'normal': self.__normal_launch,
            'manual': self.__manual_launch
        }

        try:
            return launch_types[launch_type](program_path)
        except (AppStartError, ProcessNotFoundError, MatchError):
            log.exception(ERRORS.get('file_not_found').format(file=self.program_name))

    @staticmethod
    def __normal_launch(program_path) -> Application:
        """Description in launch() doc"""

        app = Application(backend='uia').start(program_path)
        return app

    def __manual_launch(self, program_path: str) -> Application:
        """Description in launch() doc"""

        file_parse_dict = Helper.parse_file_path(program_path)
        self.launch(f"explorer.exe {file_parse_dict['folder_path']}", "normal")
        explorer = self.connect('explorer.exe')
        dlg = explorer[file_parse_dict.get('folder_name')]
        dlg[file_parse_dict['file_name']].click_input(double=True)
        sleep(5)
        explorer.kill()
        return self.connect(file_parse_dict['file_name'])

    @classmethod
    def _wait_process(cls, app, exit_comp, progress_field) -> bool:
        """
        :param app:
        :param exit_comp:
        :param progress_field:
        :return: True - if finished after the progress bar is full,
                 False - if finished after closing the pop-up window completion
        """

        dlg = app['Dialog']
        while True:
            sleep(5)
            progress_str = re.sub("[^0-9]", "", dlg[progress_field].window_text())
            progress = int(progress_str) if len(progress_str) > 0 else 0
            if progress >= 100:
                return False
            if dlg[exit_comp].exists():
                dlg.OKButton.click_input()
                return True

    def work(self):
        pass


class BtcToolsWorker(AppWorker):

    def work(self) -> bool:
        """Retrieve data on all devices in the network and save it in Excel format"""

        # scanning
        dlg = self.program_obj['Dialog']
        sleep(5)
        if dlg.NoButton:
            dlg.NoButton.click_input()
        dlg.ScanButton.click_input()
        self._wait_process(self.program_obj, 'Dialog', 'Progress')
        sleep(1)
        # export
        dlg.Header5.click_input()
        dlg.ExportButton.click_input()
        modal = self.program_obj['Dialog']
        sleep(1)
        today = Helper.get_cur_date('_dd_mm')
        modal[config['BtcTools']['data_folder']].click_input(button='left', double=True)
        modal.ComboBox0.type_keys(f'scan{today}')
        # saving
        modal.SaveButton.click_input()
        modal = self.program_obj['Dialog']
        if modal.YesButton:
            modal.YesButton.click_input()
        elif modal.OkButton:
            modal.OkButton.click_input()
        return True


class MercuryWorker(AppWorker):

    energy_meters = {
            'reset_energy': 'Static74',
            'previous_day': 'Static86'
        }

    def work(self) -> bool:
        """full process of Mercury.exe work"""

        values_dict = self.get_data()

        if values_dict and sum(values_dict.values()) > 0:
            try:
                ew = ExcelWriter()
                ew.write_workbook_data(values_dict, self.app_conf['data_path'])
            except Exception as ex:
                log.exception(f'Excel write exception: {ex}')
                np = NotepadWriter()
                np.write_data(values_dict, self.app_conf['notepad_data_path'])
            finally:
                return True
        else:
            log.exception(ERRORS.get('mercury_data_incorrect'))
            return False

    def get_data(self) -> dict:
        """function to get the values of all electricity meters"""

        mercury = self.program_obj
        meters_values = {}
        meters_list = Helper.convert_str_to_list(self.app_conf['meter_indexes'])
        for meter in meters_list:
            meters_values[meter] = self.get_meter_data(mercury, meter)
        return meters_values

    @classmethod
    def get_meter_data(cls, app: Application, meter_id: int) -> dict | None:
        """
        function to get the values of one electricity meter
        :param app: Application instance
        :param meter_id: Meter identifier
        :return: Meter value
        """

        values = {}

        dlg = app['Dialog']
        dlg['Параметры связиHyperlink'].click_input()
        sleep(1)
        dlg['СчетчикEdit'].set_text(u'')
        dlg['СчетчикEdit'].type_keys(f'{meter_id}')
        dlg['Уровень доступаEdit'].set_text(u'111111')
        dlg['\xa0Соединить\xa0'].click_input()
        if cls._wait_process(app, 'Ошибка!', 'Static3'):
            return
        dlg.Hyperlink9.click_input()
        dlg['Параметры связиRadioButton0'].click_input()
        dlg.Button1.click_input()
        sleep(5)
        for value_type in cls.energy_meters:
            value = dlg[value_type].window_text()
            try:
                values[value_type] = float(value)
            except ValueError as e:
                log.exception(f"Can't get value: {value}. Exception: {e}")
        return values


class ExcelWriter:

    # TODO: передавать в метод
    energy_meters = ['previous_day', 'reset_energy']

    def __init__(self):
        self.app_conf = config['Excel']
        self.meters_row = int(self.app_conf['meter_index_row'])
        self.meters_columns = range(int(self.app_conf['meter_index_col_first']),
                                    int(self.app_conf['meter_index_col_last']))

    @classmethod
    def check_last_date(cls, wb: openpyxl.load_workbook, sheet: str) -> bool:
        sheet = wb[sheet]
        last_row = sheet.max_row
        last_date = sheet.cell(row=last_row, column=1).value
        if last_date == Helper.get_cur_date('dd.mm.yyyy'):
            return False
        return True

    def write_workbook_data(self, data: dict, data_path: str):
        """Recording dict data to Excel row by keys"""

        self.close_workbook(config['Excel']['program_path'], config['Mercury']['data_path'])
        wb = openpyxl.load_workbook(data_path)
        for meter_type in self.energy_meters:
            self.write_sheet_data(wb, data, meter_type)
        self.save_data(wb, data_path)

    def write_sheet_data(self, wb: openpyxl.load_workbook, data: dict, meter_type):
        current_date = Helper.get_cur_date('dd.mm.yyyy')

        if meter_type not in wb.sheetnames:
            log.info(f"Sheet {meter_type} doesn't exists")
            return
        sheet = wb[meter_type]
        last_row = sheet.max_row
        if not self.check_last_date(wb, meter_type):
            log.info(f"{meter_type} data on the {current_date} exists")
            return
        date_cell = sheet.cell(row=last_row + 1, column=1)
        date_cell.value = current_date

        for column in self.meters_columns:
            self.write_cell_data(data, sheet, meter_type, column, last_row)

    def write_cell_data(self, data, sheet, meter_type, column, row):
        meter_number = sheet.cell(row=self.meters_row, column=column).value
        if not meter_number:
            log.warning(f"Incorrect meter number in excel file")
            return
        if not (meter_data := data.get(meter_number)):
            log.warning(f"Data of the meter {meter_number} doesn't exists")
            return
        current_cell = sheet.cell(row=row + 1, column=column)
        current_cell.value = meter_data.get(meter_type)

    @staticmethod
    def save_data(app, data_path: str):
        try:
            app.save(data_path)
        except PermissionError:
            current_date = Helper.get_cur_date('_dd_mm')
            app.save(data_path.replace('.xlsx', f'{current_date}.xlsx'))

    @staticmethod
    def close_workbook(program_path: str, data_path: str):
        try:
            excel = AppWorker.connect(program_path)
            data_file = Helper.get_file_name(data_path)
            if excel[data_file].exists(timeout=5):
                wb = excel[data_file]
                wb['CloseButton'].click()
                if wb['SaveButton'].exists():
                    wb['SaveButton'].click()
                sleep(3)
        except Exception as e:
            log.warning(f"Workbook closing exception: {e}")


class NotepadWriter:

    @classmethod
    def write_data(cls, data: dict, data_path: str):
        """Recording dict data to Notepad row by keys"""
        val_str = ' | '.join(map(str, data.values())).rjust(7, " ") + '\n'
        full_str = Helper.get_cur_date('dd.mm.yyyy') + ' | ' + val_str
        with open(data_path, 'a') as f:
            f.write(full_str)


class WithAppRunner:

    def __init__(self, program_name):
        self.program_name = program_name
        self.application = self.invalidate_app(self.program_name)

    def __enter__(self):
        try:
            log.info(f"{self.application.program_name} is running...")
            return self.application
        except Exception as ex:
            log.exception(f"App '{self.program_name}' start exception: {ex}")

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.application.program_obj:
            self.application.program_obj.kill()
        log.info(f'{self.application.program_name} completed')

    @staticmethod
    def invalidate_app(program_name):

        apps_dict = {
            'btctools': BtcToolsWorker,
            'mercury': MercuryWorker
        }

        if apps_dict.get(program_name):
            return apps_dict[program_name]()
        else:
            raise ModuleNotFoundError(f'{program_name} class not found')
