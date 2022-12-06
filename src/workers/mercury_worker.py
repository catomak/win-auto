from .app_worker import AppWorker, Application
from src.service import ERRORS, log, Helper, config
from time import sleep
import openpyxl


class MercuryWorker(AppWorker):

    energy_meters = {
            'reset_energy': {
                'field': 'Static86',
                'button': 'RadioButton0'
            },
            'previous_day': {
                'field': 'Static74',
                'button': 'RadioButton10'
            }
        }

    def work(self) -> bool:
        """full process of Mercury.exe work"""

        values_dict = self.get_data()

        if not values_dict:
            log.exception(ERRORS.get('mercury_data_incorrect'))
            return False

        try:
            ew = ExcelWriter()
            ew.write_workbook_data(values_dict, self.app_conf['data_path'])
        except Exception as ex:
            log.exception(f'Excel write exception: {ex}')
            np = NotepadWriter()
            np.write_data(values_dict, self.app_conf['notepad_data_path'])
        finally:
            return True

    def get_data(self) -> dict:
        """function to get the values of all electricity meters"""

        meters_values = {}
        meters_list = Helper.convert_str_to_list(self.app_conf['meter_indexes'])
        for meter in meters_list:
            meters_values[meter] = self.get_meter_data(meter)
        print(meters_values)
        return meters_values

    def get_meter_data(self, meter_id: int) -> dict | None:
        """
        function to get the values of one electricity meter
        :param app: Application instance
        :param meter_id: Meter identifier
        :return: Meter value
        """

        values = {}
        self.connect_to_meter(meter_id)

        for value_type, value_data in self.energy_meters.items():
            self.main_dlg[value_data.get('button')].click_input()
            self.main_dlg.Button1.click_input()
            sleep(3)
            value = self.main_dlg[value_data.get('field')].window_text()
            try:
                values[value_type] = float(value)
            except (ValueError, TypeError) as e:
                log.exception(f"Can't get value: {value}. Exception: {e}")

        return values

    def connect_to_meter(self, meter_id):
        self.main_dlg['Параметры связиHyperlink'].click_input()
        sleep(1)
        self.main_dlg['СчетчикEdit'].set_text(u'')
        self.main_dlg['СчетчикEdit'].type_keys(f'{meter_id}')
        if not self.main_dlg['Уровень доступаEdit'].window_text():
            self.main_dlg['Уровень доступаEdit'].set_text(u'111111')
        self.main_dlg['\xa0Соединить\xa0'].click_input()
        if self._wait_process(self.program_obj, 'Ошибка!', 'Static3'):
            return
        self.main_dlg.Hyperlink9.click_input()
        return True


class ExcelWriter:

    # TODO: передавать в метод
    energy_meters = ['previous_day', 'reset_energy']

    def __init__(self):
        self.app_conf = config['Excel']
        self.meters_row = int(self.app_conf['meter_index_row'])
        self.meters_columns = range(int(self.app_conf['meter_index_col_first']),
                                    int(self.app_conf['meter_index_col_last']))

    @classmethod
    def check_last_date(cls, wb: openpyxl.load_workbook, sheet: str) -> bool:
        sheet = wb[sheet]
        last_row = sheet.max_row
        last_date = sheet.cell(row=last_row, column=1).value
        if last_date == Helper.get_cur_date('dd.mm.yyyy'):
            return False
        return True

    def write_workbook_data(self, data: dict, data_path: str):
        """Recording dict data to Excel row by keys"""

        self.close_workbook(config['Excel']['program_path'], config['Mercury']['data_path'])
        wb = openpyxl.load_workbook(data_path)
        for meter_type in self.energy_meters:
            self.write_sheet_data(wb, data, meter_type)
        self.save_data(wb, data_path)

    def write_sheet_data(self, wb: openpyxl.load_workbook, data: dict, meter_type):
        current_date = Helper.get_cur_date('dd.mm.yyyy')

        if meter_type not in wb.sheetnames:
            log.info(f"Sheet {meter_type} doesn't exists")
            return
        sheet = wb[meter_type]
        last_row = sheet.max_row
        if not self.check_last_date(wb, meter_type):
            log.info(f"{meter_type} data on the {current_date} exists")
            return
        date_cell = sheet.cell(row=last_row + 1, column=1)
        date_cell.value = current_date

        for column in self.meters_columns:
            self.write_cell_data(data, sheet, meter_type, column, last_row)

    def write_cell_data(self, data, sheet, meter_type, column, row):
        meter_number = sheet.cell(row=self.meters_row, column=column).value
        if not meter_number:
            log.warning(f"Incorrect meter number in excel file")
            return
        if not (meter_data := data.get(meter_number)):
            log.warning(f"Data of the meter {meter_number} doesn't exists")
            return
        current_cell = sheet.cell(row=row + 1, column=column)
        current_cell.value = meter_data.get(meter_type)

    @staticmethod
    def save_data(app, data_path: str):
        try:
            app.save(data_path)
        except PermissionError:
            current_date = Helper.get_cur_date('_dd_mm')
            app.save(data_path.replace('.xlsx', f'{current_date}.xlsx'))

    @staticmethod
    def close_workbook(program_path: str, data_path: str):
        try:
            excel = AppWorker.connect(program_path)
            data_file = Helper.get_file_name(data_path)
            if excel[data_file].exists(timeout=5):
                wb = excel[data_file]
                wb['CloseButton'].click()
                if wb['SaveButton'].exists():
                    wb['SaveButton'].click()
                sleep(3)
        except Exception as e:
            log.warning(f"Workbook closing exception: {e}")


class NotepadWriter:

    @classmethod
    def write_data(cls, data: dict, data_path: str):
        """Recording dict data to Notepad row by keys"""
        val_str = ' | '.join(map(str, data.values())).rjust(7, " ") + '\n'
        full_str = Helper.get_cur_date('dd.mm.yyyy') + ' | ' + val_str
        with open(data_path, 'a') as f:
            f.write(full_str)
