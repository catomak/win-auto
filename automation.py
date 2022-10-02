from pywinauto.application import Application
from time import sleep
import openpyxl
import re
from service import Helper, config


class AppWorker:

    def __init__(self, data_path=None, data=None):
        # self.program_name = self.__class__.__name__.replace('Worker', '')
        self.data = data
        self.data_path = data_path

    @staticmethod
    def connect(program: str):
        """Method for connecting to an opened program"""
        try:
            return Application(backend='uia').connect(path=program)
        except FileNotFoundError:
            print(f'fail connection to {program}')

    @classmethod
    def launch(cls, program_path: str, launch_type: str):
        """
        Function for launch a program using one of two method
        :param program_path: the full program path
        :param launch_type: if value == 'normal' an application will be started by calling the executable file
                            if value == 'manual' an application will be started by double-click on program folder
        :return: the instance of Application class type
        """

        if launch_type == 'normal':
            return cls.__normal_launch(program_path)
        elif launch_type == 'manual':
            return cls.__manual_launch(program_path)
        else:
            raise ValueError('Incorrect launch type')

    @classmethod
    def __normal_launch(cls, program_path: str):
        """Description in launch() doc"""

        app = Application(backend='uia').start(program_path)
        return app

    @classmethod
    def __manual_launch(cls, program_path: str):
        """Description in launch() doc"""

        file_name = program_path[program_path.rfind('\\')+1:]
        folder_location = program_path[:program_path.rfind('\\'+file_name)]
        folder = folder_location[folder_location.rfind('\\')+1:]

        cls.launch(f'explorer.exe {folder_location}', 'normal')
        explorer = cls.connect('explorer.exe')
        dlg = explorer[folder]
        dlg[file_name].click_input(double=True)
        sleep(5)
        explorer.kill()

        return cls.connect(file_name)

    @classmethod
    def _click_on_item(cls, app: Application, item: str):
        app[item].click_input()

    @classmethod
    def _input_text_in_item(cls, app: Application, item: str):
        pass

    @classmethod
    def _read_text_from_item(cls, app: Application, item: str):
        pass

    @classmethod
    def _wait_process_execution(cls, app: Application):
        pass


class MercuryWorker(AppWorker):

    # def __init__(self):
    #     super.__init__()

    def work(self):
        """full process of Mercury.exe work"""

        values_dict = self.get_data()
        if len(values_dict) > 0:
            try:
                ExcelWorker(config['mercury_excel_path'], values_dict).write_data()
            except FileNotFoundError:
                NotepadWorker(config['mercury_notepad_path'], values_dict).write_data()
        else:
            return Helper.ERRORS['mercury_data_incorrect']

    def get_data(self):
        """function to get the values of all electricity meters"""

        mercury = self.launch(config['Mercury_path'], 'manual')
        meters_values = {}
        for n in config['mercury_indexes']:
            meters_values[n] = self.get_meter_data(mercury, n)
        return meters_values

    @classmethod
    def get_meter_data(cls, app, meter_id):
        """
        function to get the values of one electricity meter
        :param app: Application instance
        :param meter_id: Meter identifier
        :return: Meter value
        """
        dlg = app['Dialog']
        cls._input_text_in_item(dlg, 'Параметры связиHyperlink')
        dlg['Параметры связиHyperlink'].click_input()
        sleep(1)
        dlg['СчетчикEdit'].set_text(u'')
        dlg['СчетчикEdit'].type_keys(f'{meter_id}')
        dlg['Уровень доступаEdit'].set_text(u'111111')
        dlg['\xa0Соединить\xa0'].click_input()
        for i in range(3):
            sleep(10)
            progress_str = re.sub("[^0-9]", "", dlg.Static3.window_text())
            progress = int(progress_str) if len(progress_str) > 0 else 0
            if progress >= 100:
                break
            if dlg.OKButton.exists():
                dlg_error = app['Dialog']
                dlg_error.OKButton.click_input()
                return 0
        dlg.Hyperlink9.click_input()
        dlg['Параметры связиRadioButton0'].click_input()
        dlg.Button1.click_input()
        sleep(5)
        # От сброса - Static86, за пред. сутки - Static74
        value_data = dlg.Static74.window_text()
        try:
            return float(value_data)
        except ValueError:
            return 0


class ExcelWorker(AppWorker):
    METERS_ROW = 2

    def launch(self, program_name, launch_type):
        """Method override for working with Excel"""

        return openpyxl.load_workbook(self.data_path)

    def write_data(self):
        """Recording dict data to Excel row by keys"""

        wb = self.launch('Excel', 'normal')
        sheet = wb.active
        last_row = sheet.max_row
        last_date = sheet.cell(row=last_row, column=1).value
        current_date = Helper.get_cur_date('dd.mm.yyyy')
        if last_date == current_date:
            return
        date_cell = sheet.cell(row=last_row + 1, column=1)
        date_cell.value = current_date
        for i in range(2, 11):
            self.write_cell_data(sheet, i, last_row)
        try:
            wb.save(self.data_path)
        except PermissionError:
            wb.save(self.data_path.replace('.xlsx', '_cur_date.xlsx'))

    def write_cell_data(self, sheet, column, row):
        power_meter_number = sheet.cell(row=self.METERS_ROW, column=column).value
        current_cell = sheet.cell(row=row + 1, column=column)
        current_cell.value = self.data.get(power_meter_number)


class NotepadWorker(AppWorker):

    def write_data(self):
        """Recording dict data to Notepad row by keys"""

        val_str = Helper.get_cur_date('dd.mm.yyyy') + ' | ' + ' | '.join(map(str, self.data.values())) + '\n'
        with open(self.data_path, 'a') as f:
            f.write(val_str)


class BtcToolsWorker(AppWorker):

    def work(self):
        """Retrieve data on all devices in the network and save it in Excel format"""

        btc_tool = self.launch(config['BtcTools_path'], 'normal')
        dlg = btc_tool['Dialog']
        sleep(10)
        if dlg.NoButton:
            dlg.NoButton.click_input()
        dlg.ScanButton.click_input()
        for i in range(10):
            sleep(10)
            if dlg['Comlete'].exists():
                modal = btc_tool['Dialog']
                modal.OkButton.click_input()
                break
        sleep(1)
        dlg.Header5.click_input()
        dlg.ExportButton.click_input()
        modal = btc_tool['Dialog']
        sleep(1)
        today = Helper.get_cur_date('_dd_mm')
        modal['BTC_Tool_SCAN'].click_input(button='left', double=True)
        modal.ComboBox0.type_keys(f'scan{today}')
        modal.SaveButton.click_input()
        modal = btc_tool['Dialog']
        if modal.YesButton:
            modal.YesButton.click_input()
        elif modal.OkButton:
            modal.OkButton.click_input()
        sleep(3)
        btc_tool.kill()


# looks bad, will be redone
class AppInterface:

    def __init__(self, program):
        self.program = program

    def work(self):
        eval(f'{self.program}Worker().work()')
