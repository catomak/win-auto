from datetime import datetime
from time import sleep
import telegram
from pywinauto.application import Application, AppStartError, ProcessNotFoundError, findbestmatch
from pywinauto import Desktop, application
from json import load
import schedule
import openpyxl
import re
import os

config = {}


class Helper:
    __instance = None

    ERRORS = {
        'config_not_found': 'Configuration not found. Program directory must contain config.json',
        'program_failed': ' not found. Check program and its full path in configuration file',
        'process_not_found': ' not found. Check program and its full path and that this program in the Toolbars',
        'start_time': 'Start time must be in str format (ex.: 09:00)',
        'sleep_time': 'Sleep time must be in int seconds format between 0 and 86400 (ex.: 60)',
        'mercury_data_incorrect': 'Mercury data not found or is incorrect. The data was not recorded',
        'date_format': 'Unrecognized date format. The correct formats: ',
        'automation_failed': 'Не удалось автоматически выполнить программы: ',
        'recipients_error': 'Telegram recipients for bug report not found',
        'token_error': 'token to message sending not found'
    }

    now = datetime.now()
    td = now.today()

    def __new__(cls, *args, **kwargs):
        if cls.__instance is None:
            cls._instance = super.__new__(cls)
        return cls.__instance

    def __del__(self):
        self.__instance = None

    @staticmethod
    def get_cur_date(date_format):
        now = datetime.now()
        td = now.today()
        datetime_dict = {
            'dd.mm.yyyy': now.strftime("%d.%m.%Y"),
            '_dd_mm': f'_{str(td.day).rjust(2, "0")}_{str(td.month).rjust(2, "0")}',
            'hh:mm': f'{now.strftime("%H:%M:%S")}'
        }
        if date_format in datetime_dict.keys():
            return datetime_dict[date_format]
        else:
            raise SyntaxError(Helper.ERRORS['date_format'], {str(datetime_dict.keys())})

    @staticmethod
    def get_config(path):
        configuration = {}
        try:
            with open(path) as f:
                data = load(f)
                for d in data:
                    configuration[d] = data.get(d)
            return configuration
        except FileNotFoundError:
            raise (Helper.ERRORS['config_not_found'])


class Scheduler:

    def __init__(self, work_time: str, sleep_time: int):
        self.work_time = self.invalidate_time(work_time)
        self.sleep_time = self.invalidate_time(sleep_time)

    def __new__(cls, *args, **kwargs):
        print(f'Program runs every day at {config["start_time"]}. '
              f'To change start time, edit parameter in the [program directory] ->  "config.json".')
        return super().__new__(cls)

    @staticmethod
    def invalidate_time(time):
        if type(time) == str:
            if re.match('^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$', time) is None:
                raise ValueError(Helper.ERRORS['start_time'])
        elif type(time) == int:
            if time > 86400 or time < 0:
                raise ValueError(Helper.ERRORS['sleep_time'])
        else:
            raise TypeError(Helper.ERRORS['start_time'] + '\n' + Helper.ERRORS['sleep_time'])
        return time

    def work(self, job):
        # job()
        schedule.every().day.at(self.work_time).do(job)
        while True:
            schedule.run_pending()
            sleep(self.sleep_time)


class MainApp:
    __ROUTE_MAP = {
        1: 'cls.schedule_launch()',
        2: 'cls.test_launch()',
        3: 'cls.debug_launch()',
        4: 'exit()'
    }

    @classmethod
    def route(cls):

        global config
        config = Helper.get_config('config.json')

        mode = int(input('Choose mod (1- normal, 2 - test launch, 3 - debug launch, 4 - exit): '))

        if mode in cls.__ROUTE_MAP.keys():
            eval(cls.__ROUTE_MAP[mode])
        else:
            print('Incorrect command')

    @classmethod
    def full_job(cls):
        global config
        config = Helper.get_config('config.json')
        while True:
            cls.route()

    @classmethod
    def schedule_launch(cls):
        sc = Scheduler(config['start_time'], config['sleep_time'])
        sc.work(cls.programs_automation)

    @classmethod
    def programs_automation(cls):
        app_list = config['automation_list']
        for p in app_list:
            try:
                eval(f'{p}Worker(config["{p}_path"], None).work()')
                app_list.remove(p)
            except AppStartError:
                print(p, Helper.ERRORS['program_failed'])
            except ProcessNotFoundError:
                print(p, Helper.ERRORS['process_not_found'])
            except findbestmatch.MatchError:
                print(p, Helper.ERRORS['process_not_found'])

        if len(app_list) > 0:
            TgSender(app_list, os.getenv('TG_API')).report(config['tg_recipients'])

    @classmethod
    def test_launch(cls):
        spc = "-" * 10
        print(f'{spc}One-time launch program. Start at {Helper.get_cur_date("hh:mm")} {spc}')
        cls.programs_automation()
        print(f'{spc} Test launch ended {spc}')

    @staticmethod
    def debug_launch():
        # desk = Desktop(backend='uia')
        explorer = Application().connect(path='explorer.exe')
        explorer.fleshka['Element'].click_input(double=True)


class AppWorker:

    def __init__(self, program_path, data_path, data=None):
        self.program_path = program_path
        self.program_name = self.__class__.__name__.replace('Worker', '')
        self.program_obj = None
        self.data = data
        self.data_path = data_path
        print(f'{self.program_name} is running..')

    def __del__(self):
        if self.program_obj is not None:
            self.program_obj.kill()
        print(f'{self.program_name} successfully completed')

    def launch(self):
        app = Application(backend='uia').start(self.program_path)
        self.program_obj = app
        return app

    def finish(self):
        pass

    def work(self):
        pass

    def get_data(self):
        pass

    def write_data(self):
        pass


class MercuryWorker(AppWorker):

    def work(self):
        values_dict = self.get_data()
        if len(values_dict) > 0:
            try:
                ExcelWorker(None, config['mercury_excel_path'], values_dict).write_data()
            except FileNotFoundError:
                NotepadWorker(None, config['mercury_notepad_path'], values_dict).write_data()
        else:
            return Helper.ERRORS['mercury_data_incorrect']

    @staticmethod
    def connection(title):
        pass
        # exp = Application(backend='uia').connect(path='explorer.exe')
        # tb = exp['Taskbar']
        # tb[title].click_input()
        # sleep(5)
        # return Application(backend='uia').connect(path=config.get('Mercury_path'))

    def get_data(self):
        # mercury = self.connection('Mercury.exe - Shortcut - 1 running windowButton')
        # dlg = mercury['Mercury']
        dlg = self.launch()
        meters_values = {}
        for n in config['mercury_indexes']:
            meters_values[n] = self.get_meter_data(dlg, n)
        sleep(2)
        # вынести в отдельный метод
        return meters_values

    @staticmethod
    def get_meter_data(dlg, value):
        dlg['Параметры связиHyperlink'].click_input()
        sleep(3)
        dlg['СчетчикEdit'].set_text(u'')
        dlg['СчетчикEdit'].type_keys(f'{value}')
        dlg['Уровень доступаEdit'].set_text(u'111111')
        dlg['\xa0Соединить\xa0'].click_input()
        sleep(5)
        dlg.Hyperlink9.click_input()
        dlg['Параметры связиRadioButton0'].click_input()
        dlg.Button1.click_input()
        sleep(5)
        # dlg.print_control_identifiers()
        # От сброса - Static86, за пред. сутки - Static74
        value_data = dlg.Static74.window_text()
        try:
            return float(value_data)
        except ValueError:
            return 0


class ExcelWorker(AppWorker):
    METERS_ROW = 2

    def launch(self):
        return openpyxl.load_workbook(self.data_path)

    def write_data(self):
        wb = self.launch()
        sheet = wb.active
        last_row = sheet.max_row
        last_date = sheet.cell(row=last_row, column=1).value
        current_date = Helper.get_cur_date('dd.mm.yyyy')
        if last_date != current_date:
            date_cell = sheet.cell(row=last_row + 1, column=1)
            date_cell.value = current_date
            for i in range(2, 11):
                power_meter_number = sheet.cell(row=self.METERS_ROW, column=i).value
                current_cell = sheet.cell(row=last_row + 1, column=i)
                current_cell.value = self.data.get(power_meter_number)
            try:
                wb.save(self.data_path)
            except PermissionError:
                wb.save(self.data_path.replace('.xlsx', '_cur_date.xlsx'))


class NotepadWorker(AppWorker):

    def write_data(self):
        val_str = Helper.get_cur_date('dd.mm.yyyy') + ' | ' + ' | '.join(map(str, self.data.values())) + '\n'
        with open(self.data_path, 'a') as f:
            f.write(val_str)


class BtcToolWorker(AppWorker):

    def work(self):
        btc_tool = self.launch()
        dlg = btc_tool[config.get('btc_tool_v')]
        sleep(10)
        if 'Dialog' in str(btc_tool.windows()):
            modal = btc_tool['Dialog']
            if modal.NoButton:
                modal.NoButton.click_input()
        dlg.ScanButton.click_input()
        sleep(40)
        modal = btc_tool['Dialog']
        modal.OkButton.click_input()
        sleep(1)
        dlg.Header5.click_input()
        dlg.ExportButton.click_input()
        modal = btc_tool['Dialog']
        sleep(1)
        today = Helper.get_cur_date('_dd_mm')
        modal.BTC_Tool_SCANListItem.click_input(button='left', double=True)
        modal.ComboBox0.type_keys(f'scan{today}')
        modal.SaveButton.click_input()
        modal = btc_tool['Dialog']
        if modal.YesButton:
            modal.YesButton.click_input()
        elif modal.OkButton:
            modal.OkButton.click_input()
        sleep(3)
        btc_tool.kill()


class Sender:

    def __init__(self, programs: list, token: str):
        self.programs = programs
        self.token = self.invalidate_token(token)
        self.message = Helper.ERRORS['automation_failed'] + ', '.join(programs)

    @staticmethod
    def invalidate_token(token):
        if type(token) == str and len(token) > 0:
            return token
        else:
            raise ValueError(Helper.ERRORS['token_error'])

    def send_message(self, recipient):
        pass

    def report(self, recipients):
        pass


class TgSender(Sender):

    def send_message(self, chat_id):
        bot = telegram.Bot(self.token)
        bot.send_message(text=self.message, chat_id=chat_id)

    def report(self, recipients):
        if len(recipients) > 0:
            for recipient in recipients:
                self.send_message(recipient)
            print('Error report sent to telegram')
        else:
            print(Helper.ERRORS['recipients_error'])


def main():
    MainApp.full_job()


if __name__ == "__main__":
    main()
